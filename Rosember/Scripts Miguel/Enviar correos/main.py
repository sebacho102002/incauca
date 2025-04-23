import pandas as pd
import os
import win32com.client as win32
import tkinter as tk
from tkinter import filedialog, messagebox
import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from shutil import copyfile

def enviar_email(to_addrs, cc_addrs, subject, body, archivos):
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = to_addrs
    mail.CC = cc_addrs
    mail.Subject = subject
    mail.Body = body
    for archivo in archivos:
        mail.Attachments.Add(archivo)
    mail.Send()

def confirmar_envio():
    root = tk.Tk()
    root.withdraw()
    return messagebox.askyesno("Confirmación", "Los archivos han sido creados. ¿Desea enviar los correos?")

def seleccionar_archivo():
    root = tk.Tk()
    root.withdraw()
    ruta_predeterminada = 'C:\\Geodata\\Mapas_Despoblacion\\Envio Programa vuelo'
    archivo_seleccionado = filedialog.askopenfilename(
        initialdir=ruta_predeterminada,
        title="Seleccionar archivo Excel",
        filetypes=(("Archivos Excel", "*.xlsx"), ("Todos los archivos", "*.*"))
    )
    if not archivo_seleccionado:
        messagebox.showerror("Error", "No se seleccionó ningún archivo. El proceso fue cancelado.")
        raise FileNotFoundError("No se seleccionó ningún archivo. El proceso fue cancelado.")
    return archivo_seleccionado

def notificar_haciendas_faltantes(haciendas_faltantes):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Haciendas Faltantes", f"Las siguientes haciendas no tienen un proveedor asignado:\n{', '.join(haciendas_faltantes)}")

def mostrar_error(mensaje):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror("Error", mensaje)

def mostrar_exito(mensaje):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo("Proceso Completado", mensaje)

def guardar_con_estilo(template_path, df, output_path, sheet_name):
    copyfile(template_path, output_path)
    wb = load_workbook(output_path)
    ws = wb.active
    ws.title = sheet_name

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    for i, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for j, value in enumerate(row, start=1):
            cell = ws.cell(row=i, column=j)
            header = ws.cell(row=1, column=j).value
            if header == "Hacienda":
                cell.number_format = '@'
                cell.value = str(value).zfill(6)
            else:
                cell.value = value

    # Ajustar anchos de columnas
    for col in ws.iter_cols(min_row=1, max_row=1):
        header = col[0].value
        if header == " Nombre":
            ws.column_dimensions[col[0].column_letter].width = 30 #ojo con el espacio en la columna nombre. viene asi desde mapa de despoblacion y se trata igual
        elif header == "Zona":
            ws.column_dimensions[col[0].column_letter].width = 25

    wb.save(output_path)

def main():
    try:
        file_path = seleccionar_archivo()
        correos_path = 'C:\\Geodata\\Mapas_Despoblacion\\Envio Programa vuelo\\Correos.xlsx'
        output_folder = os.path.dirname(file_path)
        output_proveedores = os.path.join(output_folder, 'Proveedores')
        os.makedirs(output_proveedores, exist_ok=True)

        df = pd.read_excel(file_path, sheet_name='Hoja1')
        correos_df = pd.read_excel(correos_path)
        df['Ult.Cor/Siem'] = pd.to_datetime(df['Ult.Cor/Siem'], errors='coerce').dt.strftime('%d/%m/%Y')
        df_proveedores = df[df['Tenencia'] == 51]
        df = df[df['Tenencia'] != 51]
        zonas_unicas = df['Zona'].unique()

        archivos_generados = []
        for zona in zonas_unicas:
            df_filtrado = df[df['Zona'] == zona]
            output_file = os.path.join(output_folder, f'{zona.replace(" ", "_")}.xlsx')
            guardar_con_estilo(file_path, df_filtrado, output_file, zona)
            archivos_generados.append((zona, output_file))

        archivos_proveedores = {}
        proveedores_unicos = correos_df['Proveedor'].unique()
        haciendas_faltantes = []

        for proveedor in proveedores_unicos:
            correos_proveedor = correos_df[correos_df['Proveedor'] == proveedor]
            haciendas_proveedor = correos_proveedor['Hacienda'].values
            df_proveedor_filtrado = df_proveedores[df_proveedores['Hacienda'].isin(haciendas_proveedor)]

            if df_proveedor_filtrado.empty:
                haciendas_faltantes.extend(haciendas_proveedor)
                continue

            output_file = os.path.join(output_proveedores, f'Proveedores_{proveedor.replace(" ", "_")}.xlsx')
            guardar_con_estilo(file_path, df_proveedor_filtrado, output_file, proveedor)
            archivos_proveedores[proveedor] = (output_file, correos_proveedor)

        if not confirmar_envio():
            mostrar_error("El proceso fue cancelado por el usuario.")
            return

        mes_actual = datetime.datetime.now().strftime("%B")

        correos_por_zona = {
            'Oriental Incauca': ('efcastillo@incauca.com', 'dfcollazos@incauca.com', 'japenao@incauca.com'),
            'Central Incauca': ('nsinisterra@incauca.com', 'semora@incauca.com', 'japenao@incauca.com'),
            'Sur Incauca': ('hsaavedra@incauca.com', 'mizquierdo@incauca.com', 'japenao@incauca.com'),
            'Jamundí Incauca': ('h.zambrano31@gmail.com', 'jamarin@incauca.com', 'japenao@incauca.com'),
            'Occidental Incauca': ('jamorenol@incauca.com', 'arpuente@incauca.com', 'japenao@incauca.com'),
            'Norte Incauca': ('fcarbonero@incauca.com', 'lfgomez@incauca.com', 'japenao@incauca.com'),
            'Sur Occidental Incauca': ('lsaavedra@incauca.com', 'ammartinez@incauca.com', 'japenao@incauca.com')
        }

        for zona, output_file in archivos_generados:
            to_addrs_primary, to_addrs_optional, cc_addrs = correos_por_zona.get(
                zona,
                ('correo_predeterminado@ejemplo.com', '', 'copia_predeterminado@ejemplo.com')
            )
            to_addrs = to_addrs_primary
            if to_addrs_optional:
                to_addrs += f";{to_addrs_optional}"

            subject = f'Fotografía Aérea Drone 2025 {zona} - {mes_actual}'
            body = """Saludos,

Envío el listado de suertes para tomar fotografías aéreas con Drone la siguiente semana y determinar el mapa de resiembra, revisar las edades que se propone y señalar su aprobación.

Por favor indicarnos cuales son las suertes que se posponen una semana por bajo desarrollo (las plantas pequeñas no se detectan en la fotografía) o por alta maleza en el surco.

También señalar las suertes que no se realizan por los siguientes motivos: se resembraron o se resembrará en menos de una semana, programadas para renovación o la hacienda ya no pertenece al ingenio.

Los mapas de despoblación de las suertes aprobadas llegarán de 3 a 5 días después de la toma de fotografías, por favor esperar para realizar la labor de resiembra.

Gracias por la atención. Cordialmente Rosemberg Moreno  - Agricultura de precisión Incauca."""
            enviar_email(to_addrs, cc_addrs, subject, body, [output_file])

        for proveedor, (output_file, correos_proveedor) in archivos_proveedores.items():
            to_addrs = correos_proveedor.iloc[0]['Correo Principal']
            cc_addrs = ';'.join(correos_proveedor.iloc[0]['CC'].split(';')) if pd.notna(correos_proveedor.iloc[0]['CC']) else ''

            subject = f'Fotografía Aérea Drone 2025 {proveedor} - {mes_actual}'
            body = f"""Buena tarde, comparto consolidado de suertes para volar con drone para despoblación del mes {mes_actual}. GRACIAS

En Incauca estamos optimizando las labores de campo y cosecha mediante el uso de fotografías aéreas con Drone de alto detalle para determinar lo siguiente:
1. Mapa despoblación y resiembra (Distribución de paquetes)
2. Verificar área neta en caña. 
3. Líneas de surco para cosechar con piloto automático y disminuir pisoteo de la maquinaria.
4. Topografía para identificar problemas de drenaje. 
5. Mapa de malezas.

Los mapas de resiembra de las suertes aprobadas llegarán de 3 a 5 días después de la toma de fotografías. Toda esta información se mostrará en la plataforma web https://gis.manglar.com/

El costo por hectárea de esta tecnología es de $ 30.500 acordada con la empresa CNX - Manglar, Incauca asumirá el 50% resultando un costo por hectárea para el proveedor de $ 15.250."""
            enviar_email(to_addrs, cc_addrs, subject, body, [output_file])

        if haciendas_faltantes:
            notificar_haciendas_faltantes(haciendas_faltantes)

        mostrar_exito("Todos los correos se enviaron exitosamente.")

    except Exception as e:
        mostrar_error(f"Ocurrió un error: {e}") #si todo falla, NO LLAMEN A SEBASTIAN! Dejenlo descansar......

if __name__ == "__main__":
    main()
